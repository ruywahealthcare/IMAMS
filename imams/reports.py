import os
import datetime
import customtkinter as ctk
from tkinter import messagebox, filedialog
import database as db
from utils import compute_individual_status, TEST_TYPES, to_display_date


def _safe_import():
    errors = []
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl")
    try:
        from reportlab.lib.pagesizes import A4
    except ImportError:
        errors.append("reportlab")
    try:
        import pandas
    except ImportError:
        errors.append("pandas")
    return errors


def _get_report_header():
    return {
        'org': db.get_setting('org_name', 'Organisation'),
        'unit': db.get_setting('unit_name', 'HQ Unit'),
        'institute': db.get_setting('institute_name', 'RUYWA Battalion'),
        'address': db.get_setting('address', ''),
        'contact': db.get_setting('contact', ''),
    }


# ──── Excel Export ────

def export_to_excel(data: list, headers: list, title: str, filepath: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    hdr = _get_report_header()
    ws.merge_cells(f"A1:{chr(64 + len(headers))}1")
    ws['A1'] = hdr['institute'] + " - " + title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f"A2:{chr(64 + len(headers))}2")
    ws['A2'] = f"{hdr['org']} | {hdr['unit']} | Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(data, 4):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else '')

    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or '')) for cell in col if not isinstance(cell, MergedCell)),
            default=10
        )
        col_letter = next((cell.column_letter for cell in col
                           if not isinstance(cell, MergedCell)), None)
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(filepath)


def export_to_csv(data: list, headers: list, filepath: str):
    import csv
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data)


def export_to_pdf(data: list, headers: list, title: str, filepath: str):
    import hashlib as _hashlib
    _orig_md5 = _hashlib.md5
    def _md5_compat(*args, **kwargs):
        kwargs.pop('usedforsecurity', None)
        return _orig_md5(*args, **kwargs)
    _hashlib.md5 = _md5_compat

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    hdr = _get_report_header()
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                             rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=14, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, spaceAfter=8)

    elements.append(Paragraph(f"{hdr['institute']} \u2014 {title}", title_style))
    elements.append(Paragraph(
        f"{hdr['org']} | {hdr['unit']} | Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}",
        sub_style))
    elements.append(Spacer(1, 0.3*cm))

    table_data = [headers] + [[str(v) if v is not None else '' for v in row] for row in data]
    col_count = len(headers)
    page_w = landscape(A4)[0] - 2*cm
    col_w = page_w / col_count

    t = Table(table_data, colWidths=[col_w] * col_count, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3864')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF2F7')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    elements.append(t)
    doc.build(elements)


# ──── Report Generators ────

def _individuals_table_data():
    individuals = db.get_all_individuals()
    headers = ["Svc No", "Name", "Rank", "Trade", "COY", "Batch", "DOB",
               "Enrolled", "Blood", "Mobile", "Status", "Alert"]
    rows = []
    for ind in individuals:
        status = compute_individual_status(ind)
        rows.append([
            ind.get('service_number', ''), ind.get('name', ''), ind.get('rank', ''),
            ind.get('trade', ''), ind.get('coy', ''), ind.get('batch', ''),
            to_display_date(ind.get('date_of_birth', '')),
            to_display_date(ind.get('enrollment_date', '')),
            ind.get('blood_group', ''), ind.get('mobile_number', ''),
            "Completed" if status['monitoring_complete'] else "Active",
            status['overall_alert'].upper(),
        ])
    return headers, rows


def _test_report_data(test_type=None):
    conn = db.get_connection()
    c = conn.cursor()
    if test_type:
        c.execute("""SELECT t.*, i.name, i.service_number, i.coy, i.batch
                     FROM tests t JOIN individuals i ON t.individual_id=i.id
                     WHERE t.test_type=? ORDER BY i.name""", (test_type,))
    else:
        c.execute("""SELECT t.*, i.name, i.service_number, i.coy, i.batch
                     FROM tests t JOIN individuals i ON t.individual_id=i.id
                     ORDER BY i.name""")
    rows_raw = [dict(r) for r in c.fetchall()]
    conn.close()

    headers = ["Svc No", "Name", "COY", "Batch", "AY", "Test", "Attempt", "Date", "Result", "Remarks"]
    rows = [[r.get('service_number', ''), r.get('name', ''), r.get('coy', ''), r.get('batch', ''),
             f"AY{r['assessment_year']}", r['test_type'], f"#{r['attempt_number']}",
             to_display_date(r['date_conducted']), r.get('result', ''), r.get('remarks', '')]
            for r in rows_raw]
    return headers, rows


def _medical_report_data():
    conn = db.get_connection()
    c = conn.cursor()
    c.execute("""SELECT m.*, i.name, i.service_number, i.coy, i.batch
                 FROM medical_examinations m JOIN individuals i ON m.individual_id=i.id
                 ORDER BY i.name""")
    rows_raw = [dict(r) for r in c.fetchall()]
    conn.close()
    headers = ["Svc No", "Name", "COY", "Batch", "Type", "Date", "Category", "Result", "Remarks"]
    rows = [[r.get('service_number', ''), r.get('name', ''), r.get('coy', ''), r.get('batch', ''),
             r['medical_type'], to_display_date(r.get('date_conducted', '')), r.get('category', ''),
             r.get('result', ''), r.get('remarks', '')]
            for r in rows_raw]
    return headers, rows


def _counselling_report_data():
    conn = db.get_connection()
    c = conn.cursor()
    c.execute("""SELECT cs.*, i.name, i.service_number, i.coy, i.batch
                 FROM counselling_sessions cs JOIN individuals i ON cs.individual_id=i.id
                 ORDER BY i.name""")
    rows_raw = [dict(r) for r in c.fetchall()]
    conn.close()
    headers = ["Svc No", "Name", "COY", "Batch", "C#", "Date", "Counsellor", "Status", "Remarks"]
    rows = [[r.get('service_number', ''), r.get('name', ''), r.get('coy', ''), r.get('batch', ''),
             str(r['counselling_number']), to_display_date(r.get('date_conducted', '')),
             r.get('counsellor_name', ''), r.get('status', ''), r.get('remarks', '')]
            for r in rows_raw]
    return headers, rows


def _overdue_report_data():
    from alerts import get_all_alerts
    alerts = get_all_alerts()
    overdue = [a for a in alerts if a['alert'] == 'overdue']
    headers = ["Svc No", "Name", "Category", "Item", "Days Overdue", "Due Date"]
    rows = [[a['service_number'], a['name'], a['category'], a['item'],
             str(abs(a['days_left'])), a['end_date'].strftime('%d-%m-%Y')]
            for a in overdue]
    return headers, rows


def _due_report_data(days=90):
    from alerts import get_all_alerts
    alerts = get_all_alerts()
    due = [a for a in alerts if 0 <= a['days_left'] <= days]
    headers = ["Svc No", "Name", "Category", "Item", "Days Left", "Alert", "Due Date"]
    rows = [[a['service_number'], a['name'], a['category'], a['item'],
             str(a['days_left']), a['alert'].upper(), a['end_date'].strftime('%d-%m-%Y')]
            for a in due]
    return headers, rows


REPORT_TYPES = {
    "All Individuals": lambda: _individuals_table_data(),
    "Test Completion (All)": lambda: _test_report_data(),
    "Firing Report": lambda: _test_report_data("Firing"),
    "DST Report": lambda: _test_report_data("DST"),
    "BPET Report": lambda: _test_report_data("BPET"),
    "PPT Report": lambda: _test_report_data("PPT"),
    "Medical Report": lambda: _medical_report_data(),
    "Counselling Report": lambda: _counselling_report_data(),
    "Overdue Report": lambda: _overdue_report_data(),
    "Due within 90 Days": lambda: _due_report_data(90),
    "Due within 120 Days": lambda: _due_report_data(120),
    "Due within 150 Days": lambda: _due_report_data(150),
}


class ReportsPage(ctk.CTkFrame):
    def __init__(self, parent, current_user, **kwargs):
        super().__init__(parent, **kwargs)
        self.current_user = current_user
        self.configure(fg_color="transparent")
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Reports & Analytics",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(anchor="w", padx=20, pady=(20, 5))

        ctrl = ctk.CTkFrame(self)
        ctrl.pack(fill="x", padx=20, pady=10)

        ctk.CTkLabel(ctrl, text="Report Type:", width=110).pack(side="left", padx=10, pady=10)
        self.report_var = ctk.StringVar(value=list(REPORT_TYPES.keys())[0])
        ctk.CTkOptionMenu(ctrl, variable=self.report_var,
                          values=list(REPORT_TYPES.keys()), width=280).pack(side="left", padx=5)

        ctk.CTkButton(ctrl, text="Preview", width=100,
                      command=self._preview).pack(side="left", padx=10)
        ctk.CTkButton(ctrl, text="Export Excel", width=120,
                      command=lambda: self._export('excel')).pack(side="left", padx=5)
        ctk.CTkButton(ctrl, text="Export PDF", width=110,
                      command=lambda: self._export('pdf')).pack(side="left", padx=5)
        ctk.CTkButton(ctrl, text="Export CSV", width=110,
                      command=lambda: self._export('csv')).pack(side="left", padx=5)

        self.info_lbl = ctk.CTkLabel(self, text="", text_color="#27AE60")
        self.info_lbl.pack(anchor="w", padx=20)

        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=20, pady=5)

    def _get_data(self):
        fn = REPORT_TYPES[self.report_var.get()]
        return fn()

    def _preview(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        headers, rows = self._get_data()
        self.info_lbl.configure(text=f"{len(rows)} records  |  Report: {self.report_var.get()}")

        if not rows:
            ctk.CTkLabel(self.scroll, text="No data for this report.",
                         font=ctk.CTkFont(size=13)).pack(pady=30)
            return

        hf = ctk.CTkFrame(self.scroll)
        hf.pack(fill="x")
        col_w = max(80, min(160, 900 // len(headers)))
        for h in headers:
            ctk.CTkLabel(hf, text=h, font=ctk.CTkFont(weight="bold"),
                         width=col_w, anchor="w").pack(side="left", padx=3)

        for row in rows[:200]:
            rf = ctk.CTkFrame(self.scroll, fg_color="transparent")
            rf.pack(fill="x", pady=1)
            for val in row:
                ctk.CTkLabel(rf, text=str(val)[:30], width=col_w,
                             anchor="w").pack(side="left", padx=3, pady=2)

        if len(rows) > 200:
            ctk.CTkLabel(self.scroll,
                         text=f"... showing first 200 of {len(rows)} rows. Export to see all.",
                         text_color="#888").pack(pady=5)

    def _export(self, fmt: str):
        headers, rows = self._get_data()
        if not rows:
            messagebox.showinfo("Empty", "No data to export.")
            return

        report_name = self.report_var.get().replace(' ', '_')
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"AAAIS_{report_name}_{ts}"

        if fmt == 'excel':
            fp = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                              initialfile=default_name,
                                              filetypes=[("Excel", "*.xlsx")])
            if fp:
                try:
                    export_to_excel(rows, headers, self.report_var.get(), fp)
                    db.log_audit(self.current_user['username'], "Generated Report",
                                 f"{self.report_var.get()} Excel")
                    messagebox.showinfo("Exported", f"Saved:\n{fp}")
                except ImportError:
                    messagebox.showerror("Missing Library",
                        "openpyxl is not installed.\n\nRun in your IMAMS folder:\n"
                        "  venv\\Scripts\\pip install openpyxl")
                except Exception as e:
                    messagebox.showerror("Export Failed", str(e))
        elif fmt == 'pdf':
            fp = filedialog.asksaveasfilename(defaultextension=".pdf",
                                              initialfile=default_name,
                                              filetypes=[("PDF", "*.pdf")])
            if fp:
                try:
                    export_to_pdf(rows, headers, self.report_var.get(), fp)
                    db.log_audit(self.current_user['username'], "Generated Report",
                                 f"{self.report_var.get()} PDF")
                    messagebox.showinfo("Exported", f"Saved:\n{fp}")
                except ImportError:
                    messagebox.showerror("Missing Library",
                        "reportlab is not installed.\n\nRun in your IMAMS folder:\n"
                        "  venv\\Scripts\\pip install reportlab")
                except Exception as e:
                    messagebox.showerror("Export Failed", str(e))
        elif fmt == 'csv':
            fp = filedialog.asksaveasfilename(defaultextension=".csv",
                                              initialfile=default_name,
                                              filetypes=[("CSV", "*.csv")])
            if fp:
                try:
                    export_to_csv(rows, headers, fp)
                    db.log_audit(self.current_user['username'], "Generated Report",
                                 f"{self.report_var.get()} CSV")
                    messagebox.showinfo("Exported", f"Saved:\n{fp}")
                except Exception as e:
                    messagebox.showerror("Export Failed", str(e))
